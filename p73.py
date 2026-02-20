import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
# استيراد أدوات الانتظار الذكي/Import smart waiting tools
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ==============================================================================
#منطقة اضافة الروابط / Add links area
EXCEL_FILE_NAME = "project73.xlsx"
SHEET_NAME = "Sheet1"

PRODUCTS_MAP = {
    "C2": "https://www.amazon.eg/dp/B0DGHR9VG2",
    "C3": "https://www.amazon.eg/dp/B0FNCY8VKD", 
    "C4": "https://www.amazon.eg/dp/B0DWFPS6W6",
    "C5": "https://www.amazon.eg/dp/B0DZD9S5GC",
    "C6": "https://www.amazon.eg/dp/B0BS1QCFHX?th",
    "C7": "https://www.amazon.eg/dp/B075CP646D",
    "C8": "https://www.amazon.eg/dp/B0DW1X3Z4D",
    "C9": "https://www.amazon.eg/dp/B0DBHT1BT9",
    "C10": "https://www.amazon.eg/dp/B0FD9XXXGF",
    "C11": "https://www.amazon.eg/dp/B0FQFK93TR",
}
# ==============================================================================

def get_amazon_price(driver, url):
    try:
        driver.get(url)
        
        # --- الانتظار الذكي/Smart waiting ---
        # انتظر حتى يظهر عنصر السعر على الشاشة (بحد أقصى 15 ثانية)/Wait until the price item appears on the screen (maximum 15 seconds)
        wait = WebDriverWait(driver, 15)
        price_element = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "a-price-whole")))
        
        # بمجرد ظهور العنصر، نقوم بسحبه / Once the item appears, we drag it.
        price_whole = price_element.text
        price_fraction = driver.find_element(By.CLASS_NAME, "a-price-fraction").text
        
        # تنظيف السعر وتحويله لرقم / Clean the price and convert it to a number
        full_price = f"{price_whole}.{price_fraction}".replace(",", "")
        return float(full_price)
        
    except Exception as e:
        print(f"? تعذر جلب السعر للرابط {url[:30]}... (ربما بسبب ضعف الإنترنت أو تغير الرابط)")
        return None

def main():
    print("--- جاري تشغيل المتصفح الصامت بالانتظار الذكي /The silent browser is running in smart hold. ---")
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    
    try:
        wb = load_workbook(EXCEL_FILE_NAME)
        ws = wb[SHEET_NAME]
    except Exception as e:
        print(f"خطأ في فتح الملف / Error opening file: {e}")
        return

    for cell_address, url in PRODUCTS_MAP.items():
        print(f"فحص الخلية / Cell examination [{cell_address}]...")
        price = get_amazon_price(driver, url)
        
        if price:
            ws[cell_address] = price
            print(f"? تم التحديث /Updated: {price}")
        
    driver.quit()
    wb.save(EXCEL_FILE_NAME)
    print("\n--- تمت العملية بنجاح / The operation was completed successfully! ---")

if __name__ == "__main__":
    main()
